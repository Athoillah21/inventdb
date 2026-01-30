import os
import tempfile
from django.core.management.base import BaseCommand
from inventory.models import DatabaseInventory

# Load .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed


class Command(BaseCommand):
    help = 'Import database inventory from SharePoint Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Download and parse only, do not save to database'
        )
        parser.add_argument(
            '--interactive',
            action='store_true',
            help='Use interactive browser login (SSO)'
        )
        parser.add_argument(
            '--username',
            type=str,
            help='SharePoint username (email)',
            default=os.environ.get('SHAREPOINT_USERNAME', '')
        )
        parser.add_argument(
            '--password',
            type=str,
            help='SharePoint password',
            default=os.environ.get('SHAREPOINT_PASSWORD', '')
        )
        parser.add_argument(
            '--sheet',
            type=str,
            help='Excel sheet name to import (default: first/active sheet)',
            default=os.environ.get('SHAREPOINT_SHEET', None)
        )

    def handle(self, *args, **options):
        from openpyxl import load_workbook

        # SharePoint Configuration
        site_url = "https://365tsel.sharepoint.com/sites/RDBMS"
        file_path = "/sites/RDBMS/Dokumen Berbagi/PostgreSQL Files/testing_atho/testing_invent.xlsx"
        
        interactive = options['interactive']
        username = options['username']
        password = options['password']
        sheet_name = options['sheet']
        dry_run = options['dry_run']

        self.stdout.write(f'Connecting to SharePoint: {site_url}')
        
        try:
            # Choose authentication method
            if interactive:
                ctx = self._get_interactive_context(site_url)
            else:
                ctx = self._get_user_credential_context(site_url, username, password)
            
            if ctx is None:
                return
            
            # Download the file to temp location
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                temp_path = tmp_file.name
            
            self.stdout.write(f'Downloading: {file_path}')
            
            with open(temp_path, 'wb') as local_file:
                ctx.web.get_file_by_server_relative_url(file_path).download(local_file).execute_query()
            
            self.stdout.write(self.style.SUCCESS(f'Downloaded to: {temp_path}'))
            
            # Parse Excel
            wb = load_workbook(temp_path, data_only=True)
            
            # Select sheet
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.stdout.write(f'Using sheet: {sheet_name}')
                else:
                    self.stdout.write(self.style.ERROR(
                        f'Sheet "{sheet_name}" not found. Available sheets: {wb.sheetnames}'
                    ))
                    return
            else:
                ws = wb.active
                self.stdout.write(f'Using active sheet: {ws.title}')
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'Headers: {headers}')
            
            # Map column names to model fields
            column_map = {
                'No': 'no',
                'Hostname': 'hostname',
                'DB': 'db',
                'Database Status': 'database_status',
                'IP OAM': 'ip_oam',
                'IP SERVICE': 'ip_service',
                'CATEGORY DATABASE': 'category_database',
                'NOTES': 'notes',
                'PORT': 'port',
                'Version': 'version',
                'Tanggal Instalasi': 'installation_date',
                'OPERATION SYSTEM': 'operating_system',
                'APPS': 'apps',
                'Apps on Product Catalog': 'apps_on_product_catalog',
                'BUSINESS CATEGORY': 'business_category',
                'Departemen': 'departemen',
                'OWNER': 'owner',
                'PIC Operational': 'pic_operational',
                'Monitoring': 'monitoring',
                'Site': 'site',
                'Installed by': 'installed_by',
                'Size Mountpoint datadir in (GiB)': 'size_mountpoint',
                'Size Database (GiB)': 'size_database',
            }
            
            # Build header index
            header_index = {}
            for i, h in enumerate(headers):
                if h in column_map:
                    header_index[column_map[h]] = i
            
            imported = 0
            skipped = 0
            
            # Clear existing data (only if not dry run)
            if not dry_run:
                DatabaseInventory.objects.all().delete()
                self.stdout.write('Cleared existing data')
            
            # Iterate rows (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Get 'no' value
                    no_idx = header_index.get('no')
                    no_val = row[no_idx] if no_idx is not None else None
                    
                    if not no_val:
                        skipped += 1
                        continue
                    
                    # Build record data
                    data = {}
                    for field, idx in header_index.items():
                        val = row[idx] if idx < len(row) else None
                        # Handle dates (Excel might return datetime objects)
                        if field == 'installation_date' and val:
                            if hasattr(val, 'strftime'):
                                val = val.strftime('%Y-%m-%d')
                            else:
                                val = str(val)
                        elif val is not None:
                            val = str(val).strip()
                        else:
                            val = ''
                        data[field] = val
                    
                    # Convert 'no' to int
                    try:
                        data['no'] = int(float(data['no']))
                    except (ValueError, TypeError):
                        skipped += 1
                        continue
                    
                    if dry_run:
                        if imported < 5:
                            self.stdout.write(f'  Sample: {data}')
                    else:
                        DatabaseInventory.objects.create(**data)
                    
                    imported += 1
                    
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'Row {row_num} error: {e}'))
                    skipped += 1
            
            # Cleanup temp file
            os.unlink(temp_path)
            
            action = 'Parsed' if dry_run else 'Imported'
            self.stdout.write(self.style.SUCCESS(f'{action} {imported} records'))
            self.stdout.write(f'Skipped {skipped} rows')
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error: {e}'))
            import traceback
            traceback.print_exc()
    
    def _get_user_credential_context(self, site_url, username, password):
        """Authenticate with username/password"""
        from office365.runtime.auth.user_credential import UserCredential
        from office365.sharepoint.client_context import ClientContext
        
        if not username or not password:
            self.stdout.write(self.style.ERROR(
                'Missing credentials. Use --interactive for browser login, or provide '
                '--username and --password.'
            ))
            return None
        
        return ClientContext(site_url).with_credentials(
            UserCredential(username, password)
        )
    
    def _get_interactive_context(self, site_url):
        """Authenticate with device code flow using Microsoft Graph"""
        import msal
        import requests
        import tempfile
        from office365.sharepoint.client_context import ClientContext
        
        # Use "Microsoft Office" app - widely available for delegated access
        client_id = "d3590ed6-52b3-4102-aeff-aad2292ab01c"  # Microsoft Office
        
        authority = "https://login.microsoftonline.com/common"
        # Request SharePoint permission through the site
        scopes = ["https://365tsel.sharepoint.com/AllSites.Read"]
        
        self.stdout.write('Starting device code login...')
        
        # Create MSAL app
        app = msal.PublicClientApplication(
            client_id,
            authority=authority
        )
        
        # Start device code flow
        flow = app.initiate_device_flow(scopes=scopes)
        
        if "user_code" not in flow:
            self.stdout.write(self.style.ERROR(f'Failed to create device flow: {flow}'))
            return None
        
        # Show instructions to user
        self.stdout.write('')
        self.stdout.write(self.style.WARNING('=' * 60))
        self.stdout.write(self.style.WARNING('AUTHENTICATION REQUIRED'))
        self.stdout.write(self.style.WARNING('=' * 60))
        self.stdout.write(f'1. Open: {flow["verification_uri"]}')
        self.stdout.write(f'2. Enter code: {flow["user_code"]}')
        self.stdout.write(self.style.WARNING('=' * 60))
        self.stdout.write('')
        self.stdout.write('Waiting for you to complete sign-in...')
        
        # Wait for user to complete login
        result = app.acquire_token_by_device_flow(flow)
        
        if "access_token" in result:
            self.stdout.write(self.style.SUCCESS('Login successful!'))
            
            # Create context with the token
            access_token = result["access_token"]
            
            def _token_provider():
                return access_token
            
            ctx = ClientContext(site_url).with_access_token(_token_provider)
            return ctx
        else:
            error = result.get("error_description", result.get("error", "Unknown error"))
            self.stdout.write(self.style.ERROR(f'Login failed: {error}'))
            return None
            return None
